import pandas as pd
import numpy as np
import teradata as td
import openpyxl
from openpyxl.styles import Side, Alignment
import os
import shutil
import datetime
import calendar
import win32com.client as win32

LAST_REPORING_DATE = (datetime.datetime.now() -
                      datetime.timedelta(days=1)).date()

LAST_D_SHEET_MONTH = LAST_REPORING_DATE.month

FIRST_DAY_REPORTING_MONTH = datetime.date(
    LAST_REPORING_DATE.year, LAST_REPORING_DATE.month, 1)

FIRST_DAY_PREVIOUS_MONTH = datetime.date(LAST_REPORING_DATE.year, LAST_REPORING_DATE.month - 1,
                                         1) if LAST_REPORING_DATE.month > 1 else datetime.date(LAST_REPORING_DATE.year - 1, 12, 1)

LAST_DAY_PREVIOUS_MONTH = datetime.date(FIRST_DAY_PREVIOUS_MONTH.year, FIRST_DAY_PREVIOUS_MONTH.month, calendar.monthrange(
    FIRST_DAY_PREVIOUS_MONTH.year, FIRST_DAY_PREVIOUS_MONTH.month)[1])

SHEET_NAMES = ['Current month', 'Last month']

CURRENT_DATA_INFO = f'{SHEET_NAMES[0]}, data from {FIRST_DAY_REPORTING_MONTH} to {LAST_REPORING_DATE}'
PREVIOUS_DATA_INFO = f'{SHEET_NAMES[1]}, data from {FIRST_DAY_PREVIOUS_MONTH} to {LAST_DAY_PREVIOUS_MONTH}'

SAMPLE_FILE = 'sample.xlsx'

THIN_B = openpyxl.styles.Border(left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))

DOTTED_B = openpyxl.styles.Border(left=openpyxl.styles.Side(style="dotted"),
                                  right=openpyxl.styles.Side(style="dotted"),
                                  top=openpyxl.styles.Side(style="dotted"),
                                  bottom=openpyxl.styles.Side(style="dotted"))

DARK_BLUE_FONT = openpyxl.styles.Font(color='00008B', bold=True)

BLACK_FONT = openpyxl.styles.Font(color='000000', bold=False)

BLUE_FILL = openpyxl.styles.PatternFill(
    start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

WHITE_FILL = openpyxl.styles.PatternFill(
    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

DATE_FORMAT = 'yyyy-mm-dd'

ALIG_CENTER = Alignment(horizontal='center')


class EaStoreExcel():
    CREATED = 0

    def __init__(self, id, df):
        # id = R022, self = df_R022
        self.id = id
        self.file_name = f'{id}.xlsx'
        self.file_dir = f'stores\{self.file_name}'
        self.full_dir = rf'some_private_dir\{self.file_dir}'

        self.df_current = df.query(
            'Order_date >= @FIRST_DAY_REPORTING_MONTH and Data_zamówienia <= @LAST_REPORING_DATE')

        self.df_previous = df.query(
            'Order_date >= @FIRST_DAY_PREVIOUS_MONTH and Data_zamówienia <= @LAST_DAY_PREVIOUS_MONTH')

        self.month = self.df_current['Order_date'].dt.month
        self.month_id = self.month.unique()

        self.update_data()
        self.sent_via_outlook()

    def update_data(self):

        if not os.path.exists(self.file_dir):
            self.create_file()

        wb = openpyxl.load_workbook(self.file_dir)

        ws_current = wb.worksheets[0]
        ws_previous = wb.worksheets[1]

        data_cell = ws_current['B4'].value

        sheet_month = LAST_D_SHEET_MONTH
        if isinstance(data_cell, datetime.datetime):
            sheet_month = data_cell.month

        max_row = ws_current.max_row
        current_data = ws_current[f"B4:G{max_row}"]

        max_row_p = ws_previous.max_row
        current_data_p = ws_previous[f"B4:G{max_row_p}"]

        if self.month_id[0] != sheet_month:
            ws_previous.delete_rows(4, max_row_p-4)
            self.add_month(ws=ws_previous, cell_range=current_data)
            ws_current.delete_rows(4, max_row-4)

        self.new_rows(ws=ws_current, cell_range=current_data, df=self.df_current,
                      fill=WHITE_FILL, font=BLACK_FONT, border=DOTTED_B, alignment=ALIG_CENTER, b2_data=CURRENT_DATA_INFO)
        self.new_rows(ws=ws_previous, cell_range=current_data_p, df=self.df_previous,
                      fill=WHITE_FILL, font=BLACK_FONT, border=DOTTED_B, alignment=ALIG_CENTER, b2_data=PREVIOUS_DATA_INFO)

        wb.save(self.file_dir)

        return

    def add_month(self, ws, cell_range):
        for row in cell_range:
            for cell in row:
                ws[cell.coordinate].value = cell.value

        return

    def create_file(self):

        shutil.copy2(SAMPLE_FILE, self.file_dir)
        self.CREATED = 1

        return

    def new_rows(self, ws, df, fill, font, border, cell_range, alignment, b2_data):

        ws['B2'].value = b2_data

        for row in cell_range:
            for cell in row:
                cell.fill = fill
                cell.border = border
                cell.font = font
                if cell == 1:
                    cell.number_format = DATE_FORMAT
                if cell != 4 and cell != 5:
                    cell.alignment = alignment

        ws.insert_rows(idx=4, amount=df.shape[0])

        self.format_table(ws=ws, df=df,
                          fill=BLUE_FILL, font=DARK_BLUE_FONT, border=THIN_B, alignment=ALIG_CENTER)

        return

    def format_table(self, ws, df, fill, font, border, alignment):

        for r in range(df.shape[0]):
            if r >= df.shape[0]:
                break
            for c in range(df.shape[1]):
                cell = ws.cell(row=r+4, column=c+2)
                cell.value = df.iloc[r, c]
                cell.border = border
                cell.fill = fill
                cell.font = font
                if c == 1:
                    cell.number_format = DATE_FORMAT
                if c != 4 and c != 5:
                    cell.alignment = alignment

        return

    def sent_via_outlook(self):
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = f'{self.id}@store.sample.com'
        mail.BCC = 'BCC@sample.com'
        mail.Subject = f'Data up to {LAST_REPORING_DATE}'
        mail.HTMLBody = f'''Short information describing data using dates'''
        mail.Attachments.Add(self.full_dir)
        mail.Send()

        return


def get_sql_data():
    udaExec = td.UdaExec(appName="Daiwd", version="1.0", logConsole=False)
    session = udaExec.connect(
        method="method", system="system", username="user", password="pass")

    query_call = "CALL PROCEDURE_TO_ONLY_NEWEST_DATA"
    session.execute(query_call)

    query = "SELECT * FROM DATEBASE.TABLE"
    df = pd.read_sql(query, session)

    session.close()

    processed_df = df.rename(columns={'renaming':'columns'})

    processed_df = processed_df.sort_values(
        by=['Store', 'Order_date'], ascending=False)

    processed_df['Order_date'] = pd.to_datetime(
        processed_df['Order_date'], format='%Y-%m-%d')

    processed_df['Order_number'] = processed_df['Order_number'].astype(
        'int')
  
    #seding reports only if store has new data
    active_stores = processed_df.query(
        'Order_date >= @LAST_REPORING_DATE')
    active_stores = active_stores['Store'].unique()

    return processed_df, active_stores


def main():

    new_data, stores = get_sql_data()

    for store in stores:
        store_df = new_data[new_data['Store'] == store]
        EaStoreExcel(id=store, df=store_df)


if __name__ == '__main__':
    main()
